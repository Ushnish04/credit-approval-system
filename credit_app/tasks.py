import os
import logging
from celery import shared_task
from django.conf import settings

logger = logging.getLogger(__name__)


@shared_task(name='ingest_customer_data')
def ingest_customer_data():
    import openpyxl
    from credit_app.models import Customer

    file_path = os.path.join(settings.DATA_DIR, 'customer_data.xlsx')

    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        customer_id, first_name, last_name, age, phone_number, monthly_salary, approved_limit = row[:7]
        if customer_id is None:
            continue
        current_debt = row[7] if len(row) > 7 and row[7] is not None else 0

        Customer.objects.update_or_create(
            customer_id=customer_id,
            defaults={
                'first_name': first_name or '',
                'last_name': last_name or '',
                'age': age,
                'phone_number': int(phone_number),
                'monthly_salary': int(monthly_salary),
                'approved_limit': int(approved_limit),
                'current_debt': float(current_debt),
            }
        )
    return "Customers ingested"


@shared_task(name='ingest_loan_data')
def ingest_loan_data():
    import openpyxl
    import datetime
    from credit_app.models import Customer, Loan

    file_path = os.path.join(settings.DATA_DIR, 'loan_data.xlsx')

    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        customer_id, loan_id, loan_amount, tenure, interest_rate, monthly_repayment, emis_paid_on_time, start_date, end_date = row[:9]
        if customer_id is None or loan_id is None:
            continue

        try:
            customer = Customer.objects.get(customer_id=int(customer_id))
        except Customer.DoesNotExist:
            continue

        def parse_date(d):
            if d is None:
                return None
            if isinstance(d, datetime.datetime):
                return d.date()
            if isinstance(d, datetime.date):
                return d
            return None

        Loan.objects.update_or_create(
            loan_id=int(loan_id),
            defaults={
                'customer': customer,
                'loan_amount': float(loan_amount),
                'tenure': int(tenure),
                'interest_rate': float(interest_rate),
                'monthly_repayment': float(monthly_repayment),
                'emis_paid_on_time': int(emis_paid_on_time) if emis_paid_on_time else 0,
                'start_date': parse_date(start_date),
                'end_date': parse_date(end_date),
            }
        )
    return "Loans ingested"


@shared_task(name='ingest_all_data')
def ingest_all_data():
    result1 = ingest_customer_data()
    result2 = ingest_loan_data()

    # Reset sequences after data is loaded
    from django.db import connection
    with connection.cursor() as cursor:
        cursor.execute("SELECT setval('customers_customer_id_seq', (SELECT MAX(customer_id) FROM customers))")
        cursor.execute("SELECT setval('loans_loan_id_seq', (SELECT MAX(loan_id) FROM loans))")

    return f"{result1} | {result2} | Sequences reset"